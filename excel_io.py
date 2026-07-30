import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

def read_queries(filepath: str) -> dict:
    """Read queries from the Excel file grouping by 'annexures'.

    Returns:
        A dict mapping annexure name → list of dicts: [{"query": str, "source": str | None}, ...]
    """
    xl = pd.ExcelFile(filepath)
    all_queries = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet)
        
        cols_lower = {str(c).strip().lower(): c for c in df.columns}
        
        annexure_col = cols_lower.get("annexures") or cols_lower.get("annexure")
        query_col = cols_lower.get("queries") or cols_lower.get("query")
        source_col = cols_lower.get("source")
        
        if annexure_col and query_col:
            for _, row in df.iterrows():
                annexure = row.get(annexure_col)
                q = row.get(query_col)
                
                if pd.isna(q) or pd.isna(annexure):
                    continue
                    
                annexure_name = str(annexure).strip()
                src = None
                if source_col and not pd.isna(row.get(source_col)):
                    src = str(row.get(source_col)).strip()
                    
                if annexure_name not in all_queries:
                    all_queries[annexure_name] = []
                    
                all_queries[annexure_name].append({"query": str(q).strip(), "source": src})
                
    return all_queries


# ── Citation-marker extraction helpers ────────────────────────────────────────
_SOURCE_IMAGE_RE = re.compile(r'\[Source Image:\s*(.+?)\]', re.IGNORECASE)
_SOURCE_TABLE_RE = re.compile(r'\[Source Table:\s*(.+?)\]', re.IGNORECASE)


def _extract_image_path(text: str):
    """
    Return (cleaned_text, image_path_or_None).

    Strips [Source Image: ...] and [Source Table: ...] citation markers from
    the answer text so they don't appear as raw strings in Excel cells.
    Returns the first image path found (or None if there is none).
    """
    image_path = None
    m = _SOURCE_IMAGE_RE.search(text)
    if m:
        image_path = m.group(1).strip()
    # Strip both citation types from the displayed text
    text = _SOURCE_IMAGE_RE.sub("", text)
    text = _SOURCE_TABLE_RE.sub("", text)
    return text.strip(), image_path


def _image_already_at(ws, row: int, col: int) -> bool:
    """
    Return True if *ws* already has an image anchored at (row, col).

    write_answers() re-opens and re-saves the workbook on every incremental
    save (main.py calls it after every single query, for every sheet
    processed so far). Without this check, the same image gets re-embedded
    at the same cell on every subsequent save, stacking dozens of duplicate
    copies on top of each other (see: 65 copies of one image in Annexure 2).

    Handles both possible anchor representations:
    - a plain string like "C2" (freshly set via ws.add_image(img, "C2"), not
      yet normalized by openpyxl)
    - a OneCellAnchor/TwoCellAnchor object (what you get back after
      load_workbook() re-parses a saved file); its `_from.row`/`_from.col`
      are 0-indexed, so we add 1 to match openpyxl's 1-indexed row/col args.
    """
    for img in ws._images:
        anchor = img.anchor
        try:
            if isinstance(anchor, str):
                r, c = coordinate_to_tuple(anchor)
            else:
                r = anchor._from.row + 1
                c = anchor._from.col + 1
        except Exception:
            continue
        if r == row and c == col:
            return True
    return False


def _embed_image(ws, image_path: str, row: int, col: int) -> int:
    """
    Embed *image_path* into *ws* at the given (row, col) position.

    Returns the display height of the embedded image in Excel row-height units
    (approximately 0.75 × pixel height), so the caller can resize the row.
    Returns 0 on any failure (missing file, unsupported format, etc.).
    """
    image_path = os.path.normpath(image_path) if image_path else image_path
    if not image_path or not os.path.isfile(image_path):
        print(f"  [excel_io] image not found, skipping embed: {image_path!r} (resolved: {os.path.abspath(image_path) if image_path else None!r})")
        return 0
    if _image_already_at(ws, row, col):
        # Already embedded on a prior incremental save — skip to avoid duplicates.
        return 0
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage

        pil_img = PILImage.open(image_path)
        orig_w, orig_h = pil_img.size

        # Scale so the longest edge is at most 300 px
        max_px = 300
        scale = min(1.0, max_px / max(orig_w, orig_h))
        disp_w = max(1, int(orig_w * scale))
        disp_h = max(1, int(orig_h * scale))

        xl_img = XLImage(image_path)
        xl_img.width  = disp_w
        xl_img.height = disp_h

        anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(xl_img, anchor)

        # Convert pixel height → approximate Excel row height in points (1 pt ≈ 1.33 px)
        return int(disp_h / 1.33)

    except Exception as e:
        print(f"  [excel_io] image embed error: {e}")
        return 0


def write_answers(filepath: str, answers: dict, all_queries: dict):
    """Write answers back to the Excel file into separate sheets per annexure.

    ``answers`` maps annexure name → list of answer strings (positionally aligned
    with the query rows).

    Handling per answer type:
    * Plain-text answers        → written to column B.
    * Markdown table answers    → cells written below, pointer in column B.
    * [Source Image: path] tag  → image embedded in column C of the same row;
                                  tag stripped from column B text.
    * [Source Table: path] tag  → stripped silently (not written to cells).
    """
    wb = load_workbook(filepath)
    for annexure_name, qa_pairs in answers.items():
        safe_sheet_name = str(annexure_name)[:31]
        
        if safe_sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(safe_sheet_name)
        else:
            ws = wb[safe_sheet_name]
            
        if ws["A1"].value != "Query":
            ws["A1"] = "Query"
        if ws["B1"].value != "Answer":
            ws["B1"] = "Answer"
        if ws["C1"].value != "Figure":
            ws["C1"] = "Figure"

        original_queries = all_queries.get(annexure_name, [])

        total_queries = len(qa_pairs)
        # Tables start well below the last query row (+ 3 blank rows gap)
        current_table_row = total_queries + 4

        for i, answer in enumerate(qa_pairs, start=2):
            q_text = original_queries[i-2]["query"] if (i-2) < len(original_queries) else ""
            
            # Write the query in Col A
            ws.cell(row=i, column=1, value=q_text)

            # Strip citation markers; extract image path if present
            clean_answer, image_path = _extract_image_path(answer)

            lines = [line for line in clean_answer.split('\n')]
            table_lines = [line.strip() for line in lines if line.strip().startswith('|')]

            if len(table_lines) >= 3:
                # Keep non-table lines for the main answer cell
                prose_lines = [line for line in lines if not line.strip().startswith('|')]
                prose_text = "\n".join(prose_lines).strip()
                if prose_text:
                    ans_text = f"{prose_text}\n\n→ See table below (row {current_table_row})"
                else:
                    ans_text = f"→ See table below (row {current_table_row})"
                
                ws.cell(row=i, column=2, value=ans_text)

                # Write a label in col A so the reader knows which query owns this table
                ws.cell(row=current_table_row, column=1, value=f"[Query: {q_text[:50]}...]")

                for line in table_lines:
                    # Skip markdown separator lines like |---|---|
                    if re.match(r'^\|[\s\-:|]+\|[\s\-:|]*$', line):
                        continue

                    # Parse pipe-delimited cells cleanly
                    raw_cols = line.split('|')
                    # Strip leading/trailing empty elements from surrounding '|'
                    if raw_cols and not raw_cols[0].strip():
                        raw_cols = raw_cols[1:]
                    if raw_cols and not raw_cols[-1].strip():
                        raw_cols = raw_cols[:-1]

                    # Write starting at column B (col 2) to leave col A free for labels
                    for col_idx, col_val in enumerate(raw_cols, start=2):
                        ws.cell(row=current_table_row, column=col_idx, value=col_val.strip())
                    current_table_row += 1

                current_table_row += 1  # blank row between tables
            else:
                ws.cell(row=i, column=2, value=clean_answer)

            # Embed image into column C (works for both text and table answers)
            if image_path:
                row_height_pts = _embed_image(ws, image_path, row=i, col=3)
                if row_height_pts > 0:
                    ws.row_dimensions[i].height = row_height_pts

    wb.save(filepath)
    print(f"Answers written to {filepath}")


if __name__ == "__main__":
    queries = read_queries("Queries.xlsx")
    for annexure, qs in queries.items():
        print(f"Annexure: {annexure}")
        for item in qs:
            print(f"  - {item['query'][:80]}... [Source: {item['source']}]")